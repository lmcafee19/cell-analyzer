# Helper Python File containing functions to export recorded data about cells to either an excel spreadsheet or a csv file

import os
import csv
import openpyxl
import math

'''
    Exports given cell data to an excel spreadsheet
    @param filename: Name of excel file to edit or write to. Should end with extension .xls or .xlsx
    @param coordinates: dictionary indexed by cell id containing coordinates 
    @param areas: dictionary indexed by cell id containing areas
    @param headers: iterable object containing headers for each column
'''
def culture_to_excel_file(filename, coordinates, areas, time_between_frames, area_of_frame, coordinate_headers=None, area_headers=None):
    coordinates_to_excel_file(filename, coordinates, coordinate_headers, "Positions")
    area_to_excel_file(filename, areas, area_headers, "Areas")
    stats = calc_culture_cell_statistics(coordinates, areas, time_between_frames, area_of_frame)
    culture_stats_to_excel_file(filename, stats, "Culture Stats")


'''
    Exports given dictionary containing data on an individuals cell's area and coordinates to excel file
    @param filename: Name of excel file to edit or write to. Should end with extension .xls or .xlsx
    @param data: Dictionary containing data about the cell
    @param headers: iterable object containing headers for each column
'''
def individual_to_excel_file(filename, data:dict, time_between_frames, sheetname=None):
    current_row = 1
    current_col = 1

    # If filename does not end in .xls extension exit
    if not (filename.endswith(".xls") or filename.endswith(".xlsx")):
        raise Exception("File must be of type .xls or .xlsx")

    # If file already exists, create a new sheet to store data on
    if os.path.exists(f"{filename}"):
        # Open excel file for reading and writing
        wb = openpyxl.load_workbook(filename)
        # Create new sheet to use for this data
        sheet = wb.create_sheet(sheetname)

    else:
        # Otherwise create a new excel file to write to
        # Create Workbook to store all sheets and their data
        wb = openpyxl.Workbook()
        # Rename Default sheet and set it as our active one
        sheet = wb.active
        sheet.title = sheetname

    # Write Data to sheet
    # Loop through dictionary. For every key write all of its values within the same column then move onto the next
    for key, value in data.items():
        current_row = 1
        # Add Header
        sheet.cell(current_row, current_col, key)
        current_row += 1

        # Add Data
        for entry in value:
            sheet.cell(current_row, current_col, float(entry))
            current_row += 1

        current_col += 1

    # Generate Statistics using positional and area data about cells and export that to same excel sheet
    coordinates = merge(data["X Position (mm)"], data["Y Position (mm)"])
    stats = calc_individual_cell_statistics(coordinates, data['Area (mm^2)'], time_between_frames)

    # Keep Track of the column stats start on
    stats_col = current_col
    # Reset Row
    current_row = 1

    # Add Stats Headers
    sheet.cell(current_row, current_col, "Statistic")
    current_col += 1
    sheet.cell(current_row, current_col, "Value")
    current_col += 1
    current_row += 1

    # Loop Through Stats and add them to excel sheet
    for key, value in stats.items():
        # Add Header
        current_col = stats_col
        sheet.cell(current_row, current_col, key)
        current_col += 1

        # Add Data
        sheet.cell(current_row, current_col, value)
        current_row += 1

    # Save File
    wb.save(f"{filename}")


'''
    Exports given cell coordinates to an excel spreadsheet
    @param filename: Name of excel file to edit or write to. Should end with extension .xls or .xlsx
    @param data: Dictionary indexed by cell id, and containing tuples of the coordinates for each cell
    @param headers: iterable object containing headers for each column
'''
def coordinates_to_excel_file(filename, data, headers=None, sheetname=None):
    if headers is None:
        headers = []

    current_row = 1
    current_col = 1

    # If filename does not end in .xls extension exit
    if not (filename.endswith(".xls") or filename.endswith(".xlsx")):
        raise Exception("File must be of type .xls or .xlsx")

    # If file already exists, create a new sheet to store data on
    if os.path.exists(f"{filename}"):
        # Open excel file for reading and writing
        wb = openpyxl.load_workbook(filename)
        # Add Sheet to store column/row data about this iteration
        sheet = wb.create_sheet(sheetname)
    else:
        # Otherwise create a new excel file to write to
        # Create Workbook to store all sheets and their data
        wb = openpyxl.Workbook()
        # Rename default sheet created and continue
        sheet = wb.active
        sheet.title = sheetname



    # Write Data to sheet
    # Create Headers
    for i in range(0, len(headers)):
        # Cell (row, col, data) Base 1
        sheet.cell(current_row, current_col + i, headers[i])
    current_row += 1

    # Loop through all data given then extract useful info and append it
    # Adds data to new row. Argument must be iterable object
    for key, value in data.items():
        current_col = 1
        sheet.cell(current_row, current_col, key)
        current_col += 1
        for val in value:
            # Split tuples into x and y coordinate
            val = tuple(val)
            x = val[0]
            y = val[1]

            # Place x coordinate in one column and y in the next
            sheet.cell(current_row, current_col, float(x))
            current_col += 1
            sheet.cell(current_row, current_col, float(y))
            current_col += 1

        current_row += 1

    # Save File
    wb.save(f"{filename}")


'''
    Exports given cell areas to an excel spreadsheet
    @param filename: Name of excel file to edit or write to. Should end with extension .xls or .xlsx
    @param data: Dictionary indexed by cell id, and containing data about each cell
    @param headers: iterable object containing headers for each column
'''
def area_to_excel_file(filename, data, headers=None, sheetname=None):
    if headers is None:
        headers = []

    current_row = 1
    current_col = 1

    # If filename does not end in .xls extension exit
    if not (filename.endswith(".xls") or filename.endswith(".xlsx")):
        raise Exception("File must be of type .xls or .xlsx")

    # If file already exists, create a new sheet to store data on
    if os.path.exists(f"{filename}"):
        # Open excel file for reading and writing
        wb = openpyxl.load_workbook(filename)
    else:
        # Otherwise create a new excel file to write to
        # Create Workbook to store all sheets and their data
        wb = openpyxl.Workbook()

    # Add Sheet to store column/row data about this iteration
    sheet = wb.create_sheet(sheetname)

    # Write Data to sheet
    # Create Headers
    for i in range(0, len(headers)):
        # Cell (row, col, data) Base 1
        sheet.cell(current_row, current_col + i, headers[i])
    current_row += 1

    # Loop through all data given then extract useful info and append it
    # Adds data to new row. Argument must be iterable object
    for key, value in data.items():
        current_col = 1
        sheet.cell(current_row, current_col, key)
        current_col += 1
        for val in value:
            # Remove [] from string to format it correctly for Excel
            val = str(val).replace("[", "")
            val = str(val).replace("]", "")
            # Convert Area to float to avoid number stored as text
            sheet.cell(current_row, current_col, float(val))
            current_col += 1

        # Insert Excel Formula to display Total Growth the cell Underwent
        #=INDEX(B142:DR142,MATCH(TRUE,INDEX((B142:DR142<>0),0),0))
        growth_formula = f'=INDIRECT(ADDRESS({current_row}, {current_col - 1})) - INDEX(INDIRECT(ADDRESS({current_row}, 2)):INDIRECT(ADDRESS({current_row}, {current_col})),MATCH(TRUE,INDEX((INDIRECT(ADDRESS({current_row}, 2)):INDIRECT(ADDRESS({current_row}, {current_col}))<>0),0),0))'
        sheet.cell(current_row, current_col, growth_formula)
        current_col += 1
        # Insert Excel Formula to display largest amount of change between two time intervals
        change_formula = f'=_xlfn.AGGREGATE(14, 6, INDIRECT(ADDRESS({current_row}, 2)):INDIRECT(ADDRESS({current_row}, {current_col - 2}))-INDIRECT(ADDRESS({current_row}, 3)):INDIRECT(ADDRESS({current_row}, {current_col - 1})), 1)'
        sheet.cell(current_row, current_col, change_formula)
        current_col += 1
        current_row += 1

    # Save File
    wb.save(f"{filename}")


'''
    Calculates Statistics from positional data from one cell. These Stats include:
    Total Displacement: Total distance moved, Final Distance from origin, maximum distance from origin, average distance from origin,
    Maximum speed, average speed, average angle of direction, and final angle of direction between final point and origin
    @param data List of tuples/list containing x/y coordinates of given cells location
    @param area a list of all recorded areas the cell had each frame
    @param time_between_frames Time in minutes between each frame of cell growth video
    @return Dictionary containing statistics generated 
'''
def calc_individual_cell_statistics(data, areas, time_between_frames):
    # Create dictionary to hold all calculated statistics
    stats = {}
    distances = []
    origin_distances = []
    # In mm / min
    speeds = []
    # Angle in degrees between last and current point
    angle_of_direction = []
    final_angle = 0

    if data is not None:
        # Grab origin point
        origin_x = data[0][0]
        origin_y = data[0][1]
        x = 0
        y = 0

        # Loop through all positions and calculate stats between points
        for i in range(1, len(data)):
            # Grab x and y coordinates
            x = data[i][0]
            y = data[i][1]
            prevx = data[i-1][0]
            prevy = data[i-1][1]

            # Calc distance from traveled between each step
            distance = math.dist([prevx, prevy], [x, y])
            distances.append(distance)
            # Calc distance between origin and current step
            origin_distances.append(math.dist([origin_x, origin_y], [x, y]))
            # calc current speed
            speeds.append(distance/time_between_frames)
            # calc angle of direction from last point. Because of the way opencv stores coordinates
            # ((0,0) would be the top left) we need to convert the angle by subtracting 360 degrees by it
            angle = 360 - ((math.atan2(y - prevy, x - prevx) * (180 / math.pi)) % 360)
            angle_of_direction.append(angle)

            # If on final coordinate calculate angle between this and origin point
            if i == (len(data) - 1):
                final_angle = 360 - ((math.atan2(y - origin_y, x - origin_x) * (180 / math.pi)) % 360)

        # Positional Stats
        # Total Displacement (Total Distance Traveled)
        stats["Total Displacement (mm)"] = sum(distances)
        # Final Distace from Origin
        stats["Final Distance from Origin (mm)"] = math.dist([origin_x, origin_y], [x, y])
        # Maximum Distance from origin
        stats["Maximum Distance from Origin (mm)"] = max(origin_distances)
        # Average Distance from origin
        stats["Average Distance from Origin (mm)"] = sum(origin_distances)/len(origin_distances)
        # Maximum Distance Traveled in one Interval
        stats["Maximum Distance Traveled in one Interval (mm)"] = max(distances)
        # Max Speed (distance/time)
        stats["Maximum Speed (mm/min)"] = max(speeds)
        # Average Speed
        stats["Average Speed (mm/min)"] = sum(speeds)/len(speeds)
        # Average Angle of direction from origin in degrees
        stats["Average Angle of Direction from Origin (degrees)"] = sum(angle_of_direction)/len(angle_of_direction)
        # Angle of direction from origin to final point
        stats["Angle of Direction between Origin and Final Point (degrees)"] = final_angle
        # Categorize Direction of Movement
        compass_brackets = ["E", "NE", "N", "NW", "W", "SW", "S", "SE", "E"]
        compass_lookup = round(final_angle / 45)
        stats["Compass Direction Moved"] = compass_brackets[compass_lookup]

        # Area Stats
        # Max Size
        stats["Maximum Size (mm^2)"] = max(areas)
        # Min Size
        stats["Minimum Size (mm^2)"] = min(areas)
        # Average Size
        stats["Average Size (mm^2)"] = sum(areas)/len(areas)
        # Calc change in size of the cell between first and last frame
        stats["Change in Cell Size (mm^2)"] = areas[len(areas) - 1] - areas[0]
        # Calc Average Change in growth between each time interval
        change = 0
        for i in range(1, len(areas)):
            change += areas[i] - areas[i-1]
        avg_change = change/len(areas)
        stats["Average Change in Cell Size Between one Interval (mm^2)"] = avg_change
    else:
        raise Exception("Empy Data Set Given")

    return stats


'''
    Exports given cell areas to an excel spreadsheet
    @param filename: Name of excel file to edit or write to. Should end with extension .xls or .xlsx
    @param stats: Dictionary of stats gathered about a culture. Keys will be used as headers
    @param sheetname Name of the created sheet in excel document
'''
def culture_stats_to_excel_file(filename, stats, sheetname=None):
    current_row = 1
    current_col = 1

    # If filename does not end in .xls extension exit
    if not (filename.endswith(".xls") or filename.endswith(".xlsx")):
        raise Exception("File must be of type .xls or .xlsx")

    # If file already exists, create a new sheet to store data on
    if os.path.exists(f"{filename}"):
        # Open excel file for reading and writing
        wb = openpyxl.load_workbook(filename)
    else:
        # Otherwise create a new excel file to write to
        # Create Workbook to store all sheets and their data
        wb = openpyxl.Workbook()

    # Add Sheet to store column/row data about this iteration
    sheet = wb.create_sheet(sheetname)

    # Write Headers
    sheet.cell(current_row, current_col, "Statistic")
    current_col += 1
    sheet.cell(current_row, current_col, "Value")
    current_col += 1
    current_row += 1

    # Loop Through Stats and add them to excel sheet
    for key, value in stats.items():
        # Add Header
        current_col = 1
        sheet.cell(current_row, current_col, key)
        current_col += 1

        # Add Data
        sheet.cell(current_row, current_col, value)
        current_row += 1

    # Save File
    wb.save(f"{filename}")


'''
    Calculates Statistics averaging out data from all cells
    Total Displacement: Total distance moved, Final Distance from origin, maximum distance from origin, average distance from origin,
    Maximum speed, average speed, average angle of direction, and final angle of direction between final point and origin
    @param data Dictionary indexed by cell id containing a list of tuples/list containing x/y coordinates of given cells location
    @param time_between_frames Time in minutes between each frame of cell growth video
    @return Dictionary containing statistics generated 
'''
def calc_culture_cell_statistics(positional_data, area_data, time_between_frames, area_of_frame):
    # Create dictionary to hold all calculated statistics
    stats = {}
    displacements = []
    final_distances = []
    # In mm / min
    speeds = []
    # Angle in degrees between last and current point
    angle_of_direction = []
    final_sizes = []
    growth = []
    max_cell_size = 0
    max_cell_id = 0
    min_cell_size = None
    min_cell_id = 0


    for key, data in positional_data.items():
        if data is not None:
            data = tuple(data)
            # Grab origin point
            origin_x = data[0][0]
            origin_y = data[0][1]
            x = 0
            y = 0
            distances = []

            # Loop through all positions and calculate stats between points
            for i in range(1, len(data)):
                # Only do stats on non zero values as zero is simply a placeholder specifying that the cell was not tracked that frame
                if data[i][0] != 0 or data[i][1] != 0:
                    # Grab x and y coordinates
                    x = data[i][0]
                    y = data[i][1]
                    prevx = data[i-1][0]
                    prevy = data[i-1][1]


                    # Calc Distance traveled between frames
                    distance = math.dist([prevx, prevy], [x, y])
                    distances.append(distance)

                    # calc current speed
                    speeds.append(distance/time_between_frames)

                    # If on final coordinate calculate total displacement, angle and distance between this and origin point
                    # Because of the way opencv stores coordinates
                    # ((0,0) would be the top left) we need to convert the angle by subtracting 360 degrees by it
                    if i == (len(data) - 1):
                        final_angle = 360 - ((math.atan2(y - origin_y, x - origin_x) * (180 / math.pi)) % 360)
                        angle_of_direction.append(final_angle)

                        # Calc final distance from origin
                        final_distances.append(math.dist([origin_x, origin_y], [x, y]))

                        # Record total displacement
                        displacements.append(sum(distances))

    # Generate Stats based on Cell Size
    for key, value in area_data.items():
        value = list(value)

        # Check if this cell is the new largest or smallest cell in the culture
        if min(value) != 0 and (min_cell_size is None or min(value) < min_cell_size):
            min_cell_size = min(value)
            min_cell_id = key
        if max_cell_size < max(value):
            max_cell_size = max(value)
            max_cell_id = key

        # Record Final Size for the cell
        final_sizes.append(value[len(value)-1])
        # Find first non zero value within its list of areas and record the index
        start_index = value.index(next(filter(lambda x: x!=0, value)))
        # Record difference between the initial cell and its final size
        growth.append(value[len(value) - 1] - value[start_index])


    # Total Displacement (distance traveled throughout whole video)
    stats["Average Total Displacement (mm)"] = sum(displacements)/len(displacements)
    stats["Max Distance Traveled by one Cell (mm)"] = max(displacements)
    stats["Min Distance Traveled by one Cell (mm)"] = min(displacements)
    # Average Distance from origin
    stats["Average Final Distance from Origin (mm)"] = sum(final_distances)/len(final_distances)
    # Average Speed
    stats["Average Speed (mm/min)"] = sum(speeds)/len(speeds)
    # Maximum Recorded Speed
    stats["Maximum Recorded Speed (mm/min)"] = max(speeds)
    # Minimum Recorded Speed
    stats["Minimum Recorded Speed (mm/min)"] = min(speeds)
    # Angle of direction from origin to final point
    stats["Average Angle of Direction between Origin and Final Point (degrees)"] = sum(angle_of_direction)/len(angle_of_direction)
    # Categorize Direction of Movement
    compass_brackets = ["E", "NE", "N", "NW", "W", "SW", "S", "SE", "E"]
    compass_lookup = round(stats["Average Angle of Direction between Origin and Final Point (degrees)"] / 45)
    stats["Average Compass Direction Moved"] = compass_brackets[compass_lookup]
    # Calculate Final Frame's Confluency
    # Percentage of Frame the cells take up
    stats["Final Frame's Confluency (%)"] = sum(final_sizes) / area_of_frame
    # Largest Recorded Cell
    stats["Largest Cell (mm^2)"] = max_cell_size
    stats["Largest Cell's ID"] = max_cell_id
    # Smallest Recorded Cell
    stats["Smallest Cell (mm^2)"] = min_cell_size
    stats["Smallest Cell's ID"] = min_cell_id
    # Average Size of Cells
    stats["Average Final Size of Cell (mm^2)"] = sum(final_sizes) / len(final_sizes)
    # Average cell growth/shrinkage
    stats["Average Change in Cell Size (mm^2)"] = sum(growth) / len(growth)

    return stats


'''
    Exports given cell data to a comma separated value file
    @param filename: Name of File to append to or save data to. Should end with extension .csv and contain full path as necessary
    @param data: 2d iterable object containing data about each cell
    @param headers: iterable object containing headers for each column
'''
def to_csv_file(filename, data, headers=None):
    if headers is None:
        headers = []
    # If filename does not end in .xls extension exit
    if not filename.endswith(".csv"):
        raise Exception("File must be of type .csv")

    # If file already exists, append data to the end
    if os.path.exists(f"{filename}"):
        # Open csv file in append mode
        with open(filename, "a") as file:
            csvwriter = csv.writer(file)

            # Loop through data and write each row
            for row in data:
                csvwriter.writerow(row)

    # Otherwise create a new csv file to write to
    else:
        with open(filename, "w") as file:
            csvwriter = csv.writer(file)

            # Write headers
            csvwriter.writerow(headers)

            # Loop through data and write each row
            for row in data:
                csvwriter.writerow(row)


'''
    Combines two lists together into one list containing tuples of (list1[i], list2[i])
    @param list1 List of elements the same length as list2
    @param list2 List of elements the same length as list1
    @return Merged List of tuples
'''
def merge(list1, list2):
    merged_list = [(list1[i], list2[i]) for i in range(0, len(list1))]
    return merged_list

