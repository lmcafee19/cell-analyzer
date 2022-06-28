# Helper Python File containing functions to export recorded data about cells to either an excel spreadsheet or a csv file

import os
import datetime
import csv
import openpyxl


'''
    Exports given cell data to an excel spreadsheet
    @param filename: Name of excel file to edit or write to. Should end with extension .xls or .xlsx
    @param data: Dictionary indexed by cell id, and containing data about each cell
    @param headers: iterable object containing headers for each column
'''
def to_excel_file(filename, data, headers=None, sheetname=str(datetime.datetime.now())):
    if headers is None:
        headers = []

    current_row = 1
    current_col = 1

    # If filename does not end in .xls extension exit
    if not (filename.endswith(".xls") or filename.endswith(".xlsx")):
        raise Exception("File must be of type .xls or .xlsx")

    # If file already exists, create a new sheet to store data on
    if os.path.exists(f"{filename}"):
        # Open excel file for reading and writing
        wb = openpyxl.load_workbook(filename)
    else:
        # Otherwise create a new excel file to write to
        # Create Workbook to store all sheets and their data
        wb = openpyxl.Workbook()

    # Add Sheet to store column/row data about this iteration
    sheet = wb.create_sheet(sheetname)

    # Write Data to sheet
    # Create Headers
    for i in range(0, len(headers)):
        # Cell (row, col, data) Base 1
        sheet.cell(current_row, current_col + i, headers[i])
    current_row += 1

    # Loop through all data given then extract useful info and append it
    # Adds data to new row. Argument must be iterable object
    for key, value in data.items():
        current_col = 1
        sheet.cell(current_row, current_col, key)
        current_col += 1
        for val in value:
            sheet.cell(current_row, current_col, str(val))
            current_col += 1
        current_row += 1

        #sheet.append(row)

    # Save File
    wb.save(f"{filename}")


'''
    Exports given cell data to a comma separated value file
    @param filename: Name of File to append to or save data to. Should end with extension .csv and contain full path as necessary
    @param data: 2d iterable object containing data about each cell
    @param headers: iterable object containing headers for each column
'''
def to_csv_file(filename, data, headers=None):
    if headers is None:
        headers = []
    # If filename does not end in .xls extension exit
    if not filename.endswith(".csv"):
        raise Exception("File must be of type .csv")

    # If file already exists, append data to the end
    if os.path.exists(f"{filename}"):
        # Open csv file in append mode
        with open(filename, "a") as file:
            csvwriter = csv.writer(file)

            # Loop through data and write each row
            for row in data:
                csvwriter.writerow(row)

    # Otherwise create a new csv file to write to
    else:
        with open(filename, "w") as file:
            csvwriter = csv.writer(file)

            # Write headers
            csvwriter.writerow(headers)

            # Loop through data and write each row
            for row in data:
                csvwriter.writerow(row)


#to_excel_file("../data/Test.xlsx", "Cell Stuff")
#print(str([104, 103]))
